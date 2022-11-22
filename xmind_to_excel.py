# from django.test import TestCase
from xmindparser import xmind_to_dict  # 引入XMind转换函数
from openpyxl import Workbook  # 引入工作簿类
from openpyxl.styles import Side, Border
from openpyxl.styles import PatternFill, GradientFill, Alignment  # 填充样式

# import pandas as pd
# Create your tests here.
out = xmind_to_dict("D:/Python学习/xmind文档/test.xmind")  # D:\Python学习\xmind文档
# out = xmind_to_dict("C:/Users/hst/Desktop/test.xmind")
print(out)
workbook = Workbook()  # 创建工作簿
worksheet = workbook.active  # 创建工作表
worksheet.title = out[0]['topic']['title']  # 工作表名
out = out[0]['topic']['topics']  # 生成一级标题列表数据
print(out)


def change_num(num):  # 生成字符串类型用例编号，如‘001’
    num = str(num)
    if len(num) == 1:
        num = '00'+str(int(num))
        return num
    elif len(num) == 2:
        num = '0'+str(str(num))
        return num
    else:
        return num

def change_level(four_title):  # 返回优先级
    if 'makers' in four_title.keys():  # 判断XMind是否设置了优先级
        if four_title['makers'] == ['priority-1']:
            return '高'
        elif four_title['makers'] == ['priority-2']:
            return '中'
        else:
            return '低'
    elif 'makers' not in four_title.keys():
        return ' '


test_num = 1  # 行标记

table_head = ['用例编号', '业务大类', '业务小类', '测试点', '优先级', '用例名称', '操作步骤', '预期结果']
for i in range(len(table_head)):
    worksheet.cell(1, i+1, table_head[i])  # 生成表头

for one_title in out:
    for two_title in one_title['topics']:
        for three_title in two_title['topics']:
            for four_title in three_title['topics']:
                for five_title in four_title['topics']:
                    for six_title in five_title['topics']:
                        worksheet.cell(test_num+1, 1, change_num(test_num))
                        test_num += 1
                        worksheet.cell(test_num, 2, one_title['title'])
                        worksheet.cell(test_num, 3, two_title['title'])
                        worksheet.cell(test_num, 4, three_title['title'])
                        worksheet.cell(test_num, 5, change_level(four_title))
                        worksheet.cell(test_num, 6, four_title['title'])
                        worksheet.cell(test_num, 7, five_title['title'])
                        worksheet.cell(test_num, 8, six_title['title'])
print(test_num)

# 表头设置填充颜色及对齐方式
for clo in range(1, len(table_head)+1):
    worksheet.cell(1, clo).fill = PatternFill(fill_type="solid", fgColor="99ccff")  # 第一1列到第8行填充颜色
    worksheet.cell(1, clo).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # 上下左右居中
for row in range(2, test_num+1):
    for clo_other in range(1, len(table_head) + 1):
        if clo_other <= 5:
            worksheet.cell(row, clo_other).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            worksheet.cell(row, clo_other).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 调节行宽
for cell_width in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:  # 列标
    if cell_width in ['A', 'E']:
        worksheet.column_dimensions[cell_width].width = 10.0
    elif cell_width in ['B', 'C', 'D']:
        worksheet.column_dimensions[cell_width].width = 15.0
    else:
        worksheet.column_dimensions[cell_width].width = 47.0

# 调整列高
for row_height in range(1, test_num+1):
    worksheet.row_dimensions[row_height].height = 30.0

worksheet.freeze_panes = 'H2'  # 冻结窗口
workbook.save("D:/Python学习/Excel文档/test111.xlsx")
# workbook.save("C:/Users/hst/Desktop/test.xlsx")
